import openpyxl
import csv
import matplotlib.pyplot as plt
import numpy as np
import os
class c_sp_pltbar:
    def __init__(self , path) -> None:
        self.path = path
        self.start_column = 12
        self.down_upload = {'IPv4 Download':2 , 'IPv4 Upload':4 , 'IPv6 Download':3 , 'IPv6 Upload':5}
        self.ping_jitter = {'IPv4 Delay':6 , 'IPv4 Jitter':8 , 'IPv6 Delay':7 , 'IPv6 Jitter':9}

    def f_loop(self):
        self.f_find_all_file()
        for self.csv_filename in self.file_6M:
            self.csv_filepath = self.path + '\\' + self.csv_filename
            self.f_get_all_value()
            self.f_plt6m_bar()

        for self.csv_filename in self.file_10M:
            self.csv_filepath = self.path + '\\' + self.csv_filename
            self.f_get_all_value()
            self.f_plt10m_bar()
            
        for self.csv_filename in self.file_:
            self.csv_filepath = self.path + '\\' + self.csv_filename
            self.f_get_all_value() 
            self.f_plt_bar()
            self.f_delay_jitter()

    def f_find_all_file(self):
        file_list = os.listdir(self.path)
        self.file_6M = []
        self.file_10M = []
        self.file_ = []
        for file in file_list:
            if '.csv' in file:
                if '6M' in file:
                    self.file_6M.append(file)
                elif '10M' in file:
                    self.file_10M.append(file)
                else:
                    self.file_.append(file)    
        try:
            os.mkdir(self.path+'\\limit6M')
            os.mkdir(self.path+'\\limit10M')
            os.mkdir(self.path+'\\unlimited')
            os.mkdir(self.path+'\\delay_jitter')

        except:
            pass


    def f_get_all_value(self):
        csv_data = []
        f = open(self.csv_filepath , 'r' , encoding='utf-8')
        for i in csv.reader(f) : 
            csv_data.append(i)
        workbook = openpyxl.workbook.Workbook()
        worksheet = workbook.active
        for row in csv_data:
            worksheet.append(row)
        self.sheet = workbook.worksheets[0]
        del csv_data , f , workbook ,row

    def f_plt6m_bar(self):
        for k in self.down_upload:
            self.csv_filename = self.csv_filename.replace('.csv','')
            path = self.path+'\\limit6M\\' + self.csv_filename +'---'+ k
            range_5 = 0
            range5_7 = 0
            range7_9 = 0
            range9_ = 0
            #<5 5~7 7~9 >9
            
            for i in range(self.start_column , self.sheet.max_row+1):
                try:
                    value = float(self.sheet.cell(row = i , column = self.down_upload[k]).value)
                    if value < 5 :
                        range_5 += 1
                    elif value >= 5 and value < 7 :
                        range5_7 += 1
                    elif value >= 7 and value < 9 :
                        range7_9 += 1
                    elif value >= 9 :
                        range9_ += 1
                except:
                    continue
            x = [ 1 , 2 , 3 , 4 ]        
            label = [ '<5' , '5~7' , '7~9' , '>9' ]     
            h = [ range_5 , range5_7 , range7_9 , range9_ ]   
            fig = plt.figure()
            fig.set_size_inches(12,9)
            plt.title(k,fontsize=24)
            y_major_locator=plt.MultipleLocator(2)
            ax=plt.gca()
            ax.yaxis.set_major_locator(y_major_locator)
            plt.bar(x,h,tick_label=label,width=0.5)  
            plt.ylim(0,32)
            y_ticks = np.arange(0, 32, 2)
            plt.yticks(y_ticks)
            plt.get_current_fig_manager().window.state('zoomed')
            plt.savefig(fname =  path )
            plt.cla()
            plt.clf()
            plt.close('all')

    

    def f_plt10m_bar(self):
        for k in self.down_upload:
            self.csv_filename = self.csv_filename.replace('.csv','')
            path = self.path+'\\limit10M\\' + self.csv_filename +'---'+ k
            range_5 = 0
            range5_7 = 0
            range7_9 = 0
            range9_11 = 0
            range11_ = 0
            #<5 5~7 7~9 >9
            for i in range(self.start_column , self.sheet.max_row+1):
                try:
                    value = float(self.sheet.cell(row = i , column = self.down_upload[k]).value)
                    if value < 5 :
                        range_5 += 1
                    elif value >= 5 and value < 7 :
                        range5_7 += 1
                    elif value >= 7 and value < 9 :
                        range7_9 += 1
                    elif value >= 9 and value < 11 :
                        range9_11 += 1
                    elif value >= 11 :
                        range11_ += 1
                except:
                    continue
            x = [ 1 , 2 , 3 , 4 , 5 ]        
            label = [ '<5' , '5~7' , '7~9' , '9~11' , '>11' ]     
            h = [ range_5 , range5_7 , range7_9 , range9_11 , range11_]   
            fig = plt.figure()
            fig.set_size_inches(12,9)
            plt.title(k,fontsize=24)
            y_major_locator=plt.MultipleLocator(2)
            ax=plt.gca()
            ax.yaxis.set_major_locator(y_major_locator)
            plt.bar(x,h,tick_label=label,width=0.5)  
            plt.ylim(0,32)
            y_ticks = np.arange(0, 32, 2)
            plt.yticks(y_ticks)
            plt.get_current_fig_manager().window.state('zoomed')
            
            plt.savefig(fname =  path )
            plt.cla()
            plt.clf()
            plt.close('all')

    def f_plt_bar(self):
        
        for k in self.down_upload:
            self.csv_filename = self.csv_filename.replace('.csv','')
            path = self.path+'\\unlimited\\' + self.csv_filename +'---'+ k
            range0_2 = 0
            range2_4 = 0
            range4_6 = 0
            range6_8 = 0
            range8_10 = 0
            range10_12 = 0
            range12_14 = 0
            range14_16 = 0
            range16_18 = 0
            range18_ = 0
            
            for i in range(self.start_column , self.sheet.max_row+1):
                try:
                    value = float(self.sheet.cell(row = i , column = self.down_upload[k]).value)
                    if value >= 0 and value < 2 :
                        range0_2 += 1
                    elif value >= 2 and value < 4 :
                        range2_4 += 1
                    elif value >= 4 and value < 6 :
                        range4_6 += 1
                    elif value >= 6 and value < 8 :
                        range6_8 += 1
                    elif value >= 8 and value < 10 :
                        range8_10 += 1
                    elif value >= 10 and value < 12 :
                        range10_12 += 1
                    elif value >= 12 and value < 14 :
                        range12_14 += 1    
                    elif value >= 14 and value < 16 :
                        range14_16 += 1   
                    elif value >= 16 and value < 18 :
                        range16_18 += 1  
                    elif value >= 18 :
                        range18_ += 1
                except:
                    continue
            x = [1,2,3,4,5,6,7,8,9,10]        
            label = ['0~2','2~4','4~6','6~8','8~10','10~12','12~14','14~16','16~18','>18']     
            h = [ range0_2 , range2_4 , range4_6 , range6_8 , range8_10 , range10_12 , range12_14 , range14_16 , range16_18 , range18_]   
            fig = plt.figure()
            fig.set_size_inches(12,9)
            plt.title(k,fontsize=24)
            y_major_locator=plt.MultipleLocator(2)
            ax=plt.gca()
            ax.yaxis.set_major_locator(y_major_locator)
            plt.bar(x,h,tick_label=label,width=0.5)  
            plt.ylim(0,32)
            y_ticks = np.arange(0, 32, 2)
            plt.yticks(y_ticks)
            plt.get_current_fig_manager().window.state('zoomed')
            plt.savefig(fname =  path )
            plt.cla()
            plt.clf()
            plt.close('all')


    def f_delay_jitter(self):
         for k in self.ping_jitter:
            self.csv_filename = self.csv_filename.replace('.csv','')
            path = self.path+ '\\delay_jitter\\'+self.csv_filename +'---'+ k
            range_10 = 0
            range10_20 = 0
            range20_30 = 0
            range30_40 = 0
            range40_100 = 0
            range100_ = 0
            #<5 5~7 7~9 >9
            for i in range(self.start_column , self.sheet.max_row+1):
                try:
                    value = float(self.sheet.cell(row = i , column = self.ping_jitter[k]).value)
                    if value < 10 :
                        range_10 += 1
                    elif value >= 10 and value < 20 :
                        range10_20 += 1
                    elif value >= 20 and value < 30 :
                        range20_30 += 1
                    elif value >= 30 and value < 40 :
                        range30_40 += 1
                    elif value >= 40 and value < 100 :
                        range40_100 += 1
                    elif value >= 100 :
                        range100_ += 1
                except:
                    continue
            x = [ 1 , 2 , 3 , 4 , 5 , 6 ]
            label = [ '<10' , '10~20' , '20~30' , '30~40' , '40~100' , '>100']     
            h = [ range_10 , range10_20 , range20_30 , range30_40 , range40_100 , range100_ ]   
            fig = plt.figure()
            fig.set_size_inches(12,9)
            
            plt.title(k,fontsize=24)
            y_major_locator=plt.MultipleLocator(2)
            ax=plt.gca()
            ax.yaxis.set_major_locator(y_major_locator)
            plt.bar(x,h,tick_label=label,width=0.5)  
            plt.get_current_fig_manager().window.state('zoomed')
            plt.ylim(0,32)
            y_ticks = np.arange(0, 32, 2)
            plt.yticks(y_ticks)
            plt.savefig(fname =  path )
            plt.cla()
            plt.clf()
            plt.close('all')

def main ():
    path = ( r'C:\Users\jayce\Desktop\新增資料夾 (2)\新增資料夾\新增資料夾')
    sp_pltbar = c_sp_pltbar(path)
    sp_pltbar.f_loop()


if __name__ == "__main__" :
    main()